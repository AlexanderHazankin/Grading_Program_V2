# Import the random module for selecting random feedback
import random

# Import feedback_list dictionary from the feedback_list module
from feedback_list import feedback_list

# Import the openpyxl library for working with Excel files
import openpyxl

# Import the styles module for formatting Excel cells
from openpyxl import styles

# Import the column_index_from_string function for converting column name to index
from openpyxl.utils import column_index_from_string

# Load up the workbook
book = openpyxl.load_workbook("data.xlsx")

# Access the active sheet
sheet = book.active

# Define the possible grades and the corresponding score ranges
grades = [("Fail", (0, 39)),
          ("Insufficient", (40, 54)),
          ("Barely enough", (55, 59)),
          ("Enough", (60, 69)),
          ("Almost good", (70, 79)),
          ("Good", (80, 84)),
          ("Almost very good", (85, 89)),
          ("Very good", (90, 99)),
          ("Excellent", (100, 100))]

# Get the column number for the score
score_col = column_index_from_string("B")

# Iterate through the rows, starting from the second row
for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
    name = row[0]
    score = row[score_col-1]
    grade = "N/A"
    feedback = "N/A"

    if score is None:
        grade = "<-- Error! The Score is missing"
        sheet.cell(row=index, column=3).font = openpyxl.styles.Font(color='ff0000')
    elif not isinstance(score, (int, float)):
        grade = "<-- Error! The Score is not a number"
        sheet.cell(row=index, column=3).font = openpyxl.styles.Font(color='ff0000')
    elif score < 0 or score > 100:
        grade = "<-- Error! The Score is out of range 0-100"
        sheet.cell(row=index, column=3).font = openpyxl.styles.Font(color='ff0000')

    else:
        for grade_name, score_range in grades:
            if score_range[0] <= score <= score_range[1]:
                grade = grade_name
                feedback = random.choice(feedback_list[grade])
                break
    sheet.cell(row=index, column=3).value = grade
    feedback = feedback.replace("{name}", name)
    sheet.cell(row=index, column=4).value = feedback

# Save the changes and close the workbook
book.save("data.xlsx")
book.close()
